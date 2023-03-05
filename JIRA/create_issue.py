import requests
import openpyxl as xl

workbook = xl.load_workbook(r'C:\Users\Naresh.K\PycharmProjects\API-Integration\Files\Create-Issue.xlsx')
sheet = workbook['Sheet1']
project_key = sheet.cell(2, 1).value
summary = sheet.cell(2, 2).value
desc = sheet.cell(2, 3).value
issue_type = sheet.cell(2, 4).value

url = 'https://naresh-k.atlassian.net/rest/api/2/issue/'
headers = {'Content-Type': 'application/json'}
json = {
    "fields": {
        "project":
            {
                "key": project_key
            },
        "summary": summary,
        "description": desc,
        "issuetype": {
            "name": issue_type
        }
    }
}
api_key = 'ATATT3xFfGF0-qlBqwB53QklEbFEu3bbRye4_B_OLThaxv1WtOgHsIgFez87rSI-4fhClpvXq3OwFJtqJ_kgdXnv_f_F63URtepjQsS5l03oP5ngccNvkMATS3f-GR82HQ4TIxL3HTWnmWh47CkYEtLx7wfb9A4v0HQBd5w89uQQ9LiyByhPI_k=E1187AAE'
auth = ('nareshkabali.nk@gmail.com', api_key)
response = requests.post(url=url,
                         json=json,
                         headers=headers,
                         auth=auth,
                         )

issue_id = response.json()['id']
issue_key = response.json()['key']

print(f'Issue ID is {issue_id}')
print(f'Issue Key is {issue_key}')

